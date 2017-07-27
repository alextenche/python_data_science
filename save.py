from group import *

import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
# from openpyxl.cell import get_column_letter
# from openpyxl.cell import get_column_letter
from openpyxl.utils import get_column_letter


def write_brand_and_price_file(filename, data_sample):
    brand_field_index = 5
    price_field_index = 2
    new_array = []
    for record in data_sample:
        new_record = [None] * 2
        new_record[0] = record[brand_field_index]
        new_record[1] = record[price_field_index]
        new_array.append(new_record)
    write_to_file(filename, new_array)


def write_to_file(filename, data_sample):
    example = csv.writer(open(filename, 'w', encoding='utf-8'), dialect='excel')
    example.writerows(data_sample)


def write_min_max_csv(filename, data_sample):
    min = find_max_min(data_sample, 2, "min")
    max = find_max_min(data_sample, 2, "max")
    new_array = []
    for record in data_sample:
        if (float(record[2] == min)) or (float(record[2] == max)):
            new_array.append(record)
    write_to_file(filename, new_array)


def write_two_cols(filename, data_sample, col1, col2):
    new_array = []
    for record in data_sample:
        new_record = [None] * 2
        new_record[0] = record[col1]
        new_record[1] = record[col2]
        new_array.append(new_record)
    write_to_file(filename, new_array)


def save_spreadsheet(filename, data_sample):
    wb = Workbook()
    ws = wb.active
    row_index = 1
    for rows in data_sample:
        col_index = 1
        for field in rows:
            col_letter = get_column_letter(col_index)
            ws.cell('{}{}'.format(col_letter, row_index)).value = field
            col_index += 1
        row_index += 1
    wb.save(filename)


# write_to_file('test.csv', silk_ties)
# write_brand_and_price_file('silk_ties.csv', silk_ties)
# write_min_max_csv('min_max.csv', gucci_ties[1:])
# write_two_cols('test_silk_ties', silk_ties, 2, 3)
kiton_ties = filter_col_by_string(data_from_csv, 'brandName', 'Kiton')
save_spreadsheet('kiton_ties.xls', kiton_ties)
